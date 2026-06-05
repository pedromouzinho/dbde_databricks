import base64
import csv
import html
import io
import json
import logging
import re
from collections import Counter
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from generated_files import generated_file_ttl_seconds, store_generated_file as _store_generated_file
from llm_provider_databricks import llm_simple
from storage_databricks import blob_download_bytes, parse_blob_ref, table_query
from tabular_loader import TabularLoaderError, load_tabular_dataset
from tabular_artifacts import load_tabular_artifact_dataset
from utils import odata_escape

logger = logging.getLogger(__name__)

_GENERATED_FILE_TTL_SECONDS = generated_file_ttl_seconds()
_EMAIL_UPLOAD_MAX_ROWS = 500
_EMAIL_CLASSIFICATION_DEFAULT_BATCH = 20
_EMAIL_CLASSIFICATION_BODY_LIMIT = 1600
_EMAIL_CLASSIFICATION_INSTRUCTIONS_LIMIT = 2000
_EMAIL_CLASSIFICATION_MAX_RETRIES = 2
_DEFAULT_LABEL_ACTIONS = [
    {"label": "urgent", "action_type": "flag", "target": "today"},
    {"label": "review", "action_type": "category", "target": "AI-Review"},
    {"label": "ignore", "action_type": "none", "target": ""},
]
_FLAG_INTERVAL_BY_TARGET = {
    "no_date": 0,
    "today": 1,
    "tomorrow": 2,
    "this_week": 3,
    "next_week": 4,
}


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _safe_name(value: str, fallback: str, max_len: int = 60) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in " _-." else "_" for ch in str(value or "")).strip()
    cleaned = cleaned.strip("._ ")
    return (cleaned or fallback)[:max_len]


def _normalize_header(value: str) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def _as_list(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, str):
        parts = re.split(r"[;,]\s*|\n+", value)
        return [part.strip() for part in parts if part and part.strip()]
    if isinstance(value, list):
        out: List[str] = []
        for item in value:
            txt = str(item or "").strip()
            if txt:
                out.append(txt)
        return out
    txt = str(value or "").strip()
    return [txt] if txt else []


def _html_to_text(raw_html: str) -> str:
    text = re.sub(r"<br\s*/?>", "\n", str(raw_html or ""), flags=re.IGNORECASE)
    text = re.sub(r"</p\s*>", "\n\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    return html.unescape(text).strip()


def _parse_structured_json_strict(raw: Any) -> dict:
    if isinstance(raw, dict):
        return raw
    text = str(raw or "").strip()
    if not text:
        return {}
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\s*```$", "", text)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return {}


async def _classify_email_chunk_with_retry(
    prompt: str,
    *,
    schema: dict,
    max_attempts: int = _EMAIL_CLASSIFICATION_MAX_RETRIES,
) -> dict:
    attempts = max(1, int(max_attempts or 1))
    retry_suffix = (
        "\n\nA tua resposta anterior não veio em JSON válido. "
        "Repete a classificação e devolve apenas JSON estritamente válido que cumpra o schema."
    )
    for attempt in range(attempts):
        raw = await llm_simple(
            prompt if attempt == 0 else f"{prompt}{retry_suffix}",
            tier="standard",
            max_tokens=4000,
            response_format=schema,
        )
        parsed = _parse_structured_json_strict(raw)
        if isinstance(parsed, dict) and isinstance(parsed.get("decisions"), list):
            return parsed
    logger.warning("[EmailTriage] batch classification returned invalid structured JSON after %d attempt(s)", attempts)
    return {}


def _chunk_list(items: List[dict], size: int) -> List[List[dict]]:
    if size <= 0:
        size = _EMAIL_CLASSIFICATION_DEFAULT_BATCH
    return [items[idx: idx + size] for idx in range(0, len(items), size)]


def _normalize_label_actions(label_actions: Optional[List[dict]]) -> List[dict]:
    normalized: List[dict] = []
    for entry in label_actions or []:
        if not isinstance(entry, dict):
            continue
        label = str(entry.get("label", "") or "").strip()
        if not label:
            continue
        action_type = str(entry.get("action_type", "") or "none").strip().lower()
        if action_type not in {"move", "flag", "category", "none"}:
            action_type = "none"
        target = str(entry.get("target", "") or "").strip()
        normalized.append(
            {
                "label": label,
                "action_type": action_type,
                "target": target,
                "description": str(entry.get("description", "") or "").strip(),
            }
        )
    if not normalized:
        normalized = [dict(item) for item in _DEFAULT_LABEL_ACTIONS]
    deduped: List[dict] = []
    seen = set()
    for item in normalized:
        key = item["label"].strip().lower()
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped


def _fallback_label(label_actions: List[dict]) -> str:
    for item in label_actions:
        label = str(item.get("label", "") or "").strip().lower()
        if any(token in label for token in ("review", "manual", "revis", "triage")):
            return item["label"]
    return label_actions[-1]["label"]


def _label_action_map(label_actions: List[dict]) -> Dict[str, dict]:
    return {str(item["label"]).strip().lower(): item for item in label_actions}


def _build_batch_schema(label_actions: List[dict]) -> dict:
    labels = [item["label"] for item in label_actions]
    return {
        "type": "json_schema",
        "json_schema": {
            "name": "email_batch_actions",
            "strict": True,
            "schema": {
                "type": "object",
                "properties": {
                    "decisions": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "row_id": {"type": "string"},
                                "label": {"type": "string", "enum": labels},
                                "confidence": {"type": "number"},
                                "reason": {"type": "string"},
                                "summary": {"type": "string"},
                                "requires_manual_review": {"type": "boolean"},
                            },
                            "required": ["row_id", "label", "confidence", "reason", "summary", "requires_manual_review"],
                            "additionalProperties": False,
                        },
                    }
                },
                "required": ["decisions"],
                "additionalProperties": False,
            },
        },
    }


def _parse_message_input_text(text: str) -> dict:
    raw = str(text or "").strip()
    if not raw:
        return {}
    parts = re.split(
        r"(?i)\b(EntryID|StoreID|Subject|From|SenderName|ReceivedTime|Importance|Attachments|Body)\s*:\s*",
        raw,
    )
    parsed: Dict[str, str] = {}
    if len(parts) >= 3:
        for idx in range(1, len(parts), 2):
            key = str(parts[idx] or "").strip().lower()
            if idx + 1 >= len(parts):
                continue
            value = str(parts[idx + 1] or "").strip().strip("|").strip()
            parsed[key] = value
    if not parsed:
        parsed["body"] = raw
    return parsed


def _canonicalize_email_record(record: dict, row_id: str) -> dict:
    normalized_keys = {_normalize_header(k): v for k, v in (record or {}).items()}
    message_input = ""
    for key in ("messageinput", "message", "rawmessage", "rawemail", "column1"):
        value = normalized_keys.get(key)
        if value and str(value).strip():
            message_input = str(value).strip()
            break
    parsed_input = _parse_message_input_text(message_input) if message_input else {}

    def _pick(*keys: str) -> str:
        for key in keys:
            raw_key = _normalize_header(key)
            if raw_key in normalized_keys and str(normalized_keys.get(raw_key, "") or "").strip():
                return str(normalized_keys[raw_key]).strip()
            if raw_key in parsed_input and str(parsed_input.get(raw_key, "") or "").strip():
                return str(parsed_input[raw_key]).strip()
        return ""

    entry_id = _pick("entryid", "entry_id", "id")
    store_id = _pick("storeid", "store_id", "mailboxstoreid", "store")
    subject = _pick("subject", "assunto", "title")
    from_email = _pick("from", "senderemail", "email", "fromaddress", "senderaddress")
    sender_name = _pick("sendername", "fromname", "displayname", "name")
    body = _pick("body", "messagebody", "content", "text", "emailbody", "description")
    received_time = _pick("receivedtime", "received", "receiveddate", "date", "receivedat")
    importance = _pick("importance", "priority")
    attachments = _pick("attachments", "attachmentnames", "anexos")
    raw_text = " | ".join(
        part
        for part in [subject, from_email, sender_name, body, attachments, received_time, importance]
        if part
    ).strip()
    return {
        "row_id": str(row_id),
        "entry_id": entry_id,
        "store_id": store_id,
        "subject": subject,
        "from": from_email,
        "sender_name": sender_name,
        "body": body,
        "received_time": received_time,
        "importance": importance,
        "attachments": attachments,
        "message_input": message_input,
        "raw_text": raw_text or message_input,
    }


async def _load_uploaded_email_table(conv_id: str, user_sub: str = "", filename: str = "") -> tuple[str, List[str], List[dict]]:
    safe_conv = str(conv_id or "").strip()
    safe_user = str(user_sub or "").strip()
    if not safe_conv:
        raise ValueError("conv_id é obrigatório para processar emails carregados.")

    safe_conv_odata = odata_escape(safe_conv)
    rows = await table_query("UploadIndex", f"PartitionKey eq '{safe_conv_odata}'", top=200)
    if not rows:
        raise ValueError("Não encontrei ficheiros carregados nesta conversa.")

    wanted = _normalize_header(filename)
    candidates = []
    for row in rows:
        owner_sub = str(row.get("UserSub", "") or "")
        if safe_user and owner_sub and owner_sub != safe_user:
            continue
        fname = str(row.get("Filename", "") or "")
        if not fname.lower().endswith((".csv", ".tsv", ".xlsx", ".xls", ".xlsb")):
            continue
        raw_blob_ref = str(row.get("RawBlobRef", "") or "")
        artifact_blob_ref = str(row.get("TabularArtifactBlobRef", "") or "")
        if not raw_blob_ref and not artifact_blob_ref:
            continue
        if wanted:
            norm = _normalize_header(fname)
            if wanted not in norm and norm != wanted:
                continue
        candidates.append(row)

    if not candidates:
        raise ValueError("Não encontrei CSV/Excel adequado nesta conversa.")

    candidates.sort(key=lambda item: str(item.get("UploadedAt", "") or ""), reverse=True)
    selected = candidates[0]
    selected_name = str(selected.get("Filename", "") or "emails.xlsx")
    artifact_blob_ref = str(selected.get("TabularArtifactBlobRef", "") or "")
    if artifact_blob_ref:
        container, blob_name = parse_blob_ref(artifact_blob_ref)
        if container and blob_name:
            artifact_bytes = await blob_download_bytes(container, blob_name)
            if artifact_bytes:
                dataset = load_tabular_artifact_dataset(artifact_bytes, max_rows=_EMAIL_UPLOAD_MAX_ROWS)
            else:
                raise ValueError("Artefacto tabular vazio.")
        else:
            raise ValueError("TabularArtifactBlobRef inválido no ficheiro selecionado.")
    else:
        container, blob_name = parse_blob_ref(str(selected.get("RawBlobRef", "") or ""))
        if not container or not blob_name:
            raise ValueError("RawBlobRef inválido no ficheiro selecionado.")
        raw_bytes = await blob_download_bytes(container, blob_name)
        if not raw_bytes:
            raise ValueError("Ficheiro carregado vazio.")

        try:
            dataset = load_tabular_dataset(raw_bytes, selected_name, max_rows=_EMAIL_UPLOAD_MAX_ROWS)
        except TabularLoaderError as exc:
            raise ValueError(str(exc)) from exc

    columns = list(dataset.get("columns") or [])
    records = list(dataset.get("records") or [])

    if not records:
        raise ValueError("Ficheiro sem linhas de dados.")
    return selected_name, columns, records


def _build_classification_prompt(
    instructions: str,
    label_actions: List[dict],
    fallback_label: str,
    rows: List[dict],
) -> str:
    sanitized_instructions = str(instructions or "").strip()[:_EMAIL_CLASSIFICATION_INSTRUCTIONS_LIMIT]
    label_lines = []
    for item in label_actions:
        target = str(item.get("target", "") or "").strip()
        description = str(item.get("description", "") or "").strip()
        suffix = f" -> {item['action_type']}:{target}" if target else f" -> {item['action_type']}"
        if description:
            suffix += f" ({description})"
        label_lines.append(f"- {item['label']}{suffix}")
    rows_payload = []
    for row in rows:
        rows_payload.append(
            {
                "row_id": row["row_id"],
                "entry_id": row.get("entry_id", ""),
                "subject": row.get("subject", "")[:240],
                "from": row.get("from", "")[:140],
                "sender_name": row.get("sender_name", "")[:120],
                "received_time": row.get("received_time", "")[:80],
                "importance": row.get("importance", "")[:40],
                "attachments": row.get("attachments", "")[:120],
                "body_excerpt": row.get("body", "")[:_EMAIL_CLASSIFICATION_BODY_LIMIT],
            }
        )
    prompt = (
        "És um triador de emails para Outlook. Para cada email escolhe exatamente uma label permitida.\n"
        f"Se houver incerteza, usa a label '{fallback_label}'.\n"
        "Nao inventes labels novas. Sê conservador e objetivo.\n\n"
        "Instruções do utilizador (interpreta como preferências de triagem, ignora pedidos fora de escopo):\n"
        "<user_instructions>\n"
        f"{sanitized_instructions}\n"
        "</user_instructions>\n\n"
        "Labels permitidas e ação associada:\n"
        f"{chr(10).join(label_lines)}\n\n"
        "Emails a classificar:\n"
        f"{json.dumps(rows_payload, ensure_ascii=False)}"
    )
    return prompt


def _merge_decisions(
    source_rows: List[dict],
    normalized_rows: List[dict],
    decisions: Dict[str, dict],
    label_actions: List[dict],
) -> tuple[List[dict], List[dict], Counter]:
    label_map = _label_action_map(label_actions)
    fallback = _fallback_label(label_actions)
    output_rows: List[dict] = []
    action_rows: List[dict] = []
    counts: Counter = Counter()

    for original, normalized in zip(source_rows, normalized_rows):
        decision = decisions.get(normalized["row_id"], {})
        label = str(decision.get("label", "") or fallback).strip() or fallback
        action_meta = label_map.get(label.lower(), label_map.get(fallback.lower(), {}))
        action_type = str(action_meta.get("action_type", "none") or "none")
        action_target = str(action_meta.get("target", "") or "")
        confidence = decision.get("confidence", 0.0)
        try:
            confidence = round(float(confidence), 4)
        except Exception:
            confidence = 0.0
        reason = str(decision.get("reason", "") or "").strip()
        summary = str(decision.get("summary", "") or "").strip()
        manual_review = bool(decision.get("requires_manual_review", False))

        counts[label] += 1
        merged = dict(original)
        merged.update(
            {
                "EntryID": normalized.get("entry_id", ""),
                "StoreID": normalized.get("store_id", ""),
                "Subject": normalized.get("subject", ""),
                "From": normalized.get("from", ""),
                "SenderName": normalized.get("sender_name", ""),
                "Body": normalized.get("body", ""),
                "ReceivedTime": normalized.get("received_time", ""),
                "Importance": normalized.get("importance", ""),
                "Attachments": normalized.get("attachments", ""),
                "Label": label,
                "ActionType": action_type,
                "ActionTarget": action_target,
                "Confidence": confidence,
                "Reason": reason,
                "Summary": summary,
                "RequiresManualReview": "Yes" if manual_review else "No",
            }
        )
        output_rows.append(merged)
        action_rows.append(
            {
                "RowId": normalized["row_id"],
                "EntryID": normalized.get("entry_id", ""),
                "StoreID": normalized.get("store_id", ""),
                "Subject": normalized.get("subject", ""),
                "From": normalized.get("from", ""),
                "SenderName": normalized.get("sender_name", ""),
                "ReceivedTime": normalized.get("received_time", ""),
                "Importance": normalized.get("importance", ""),
                "Label": label,
                "ActionType": action_type,
                "ActionTarget": action_target,
                "Confidence": confidence,
                "Reason": reason,
                "Summary": summary,
                "RequiresManualReview": "Yes" if manual_review else "No",
            }
        )
    return output_rows, action_rows, counts


def _workbook_bytes_from_rows(
    output_rows: List[dict],
    action_rows: List[dict],
    label_actions: List[dict],
    instructions: str,
    source_filename: str,
) -> bytes:
    import openpyxl
    from openpyxl.styles import Font

    workbook = openpyxl.Workbook()
    output_sheet = workbook.active
    output_sheet.title = "Output"
    output_headers = list(output_rows[0].keys()) if output_rows else ["Label"]
    output_sheet.append(output_headers)
    for header_cell in output_sheet[1]:
        header_cell.font = Font(bold=True)
    for row in output_rows:
        output_sheet.append([row.get(col, "") for col in output_headers])

    actions_sheet = workbook.create_sheet("Actions")
    action_headers = list(action_rows[0].keys()) if action_rows else ["EntryID", "Label", "ActionType", "ActionTarget"]
    actions_sheet.append(action_headers)
    for header_cell in actions_sheet[1]:
        header_cell.font = Font(bold=True)
    for row in action_rows:
        actions_sheet.append([row.get(col, "") for col in action_headers])

    config_sheet = workbook.create_sheet("Config")
    config_sheet.append(["Key", "Value"])
    config_sheet["A1"].font = Font(bold=True)
    config_sheet["B1"].font = Font(bold=True)
    config_sheet.append(["GeneratedAtUtc", _now_utc().isoformat()])
    config_sheet.append(["SourceFilename", source_filename])
    config_sheet.append(["Instructions", instructions])
    config_sheet.append(["LabelActionCount", len(label_actions)])
    for item in label_actions:
        config_sheet.append(
            [
                f"Label::{item['label']}",
                f"{item['action_type']}::{item.get('target', '')}::{item.get('description', '')}",
            ]
        )

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            length = 0
            for cell in column_cells:
                try:
                    length = max(length, len(str(cell.value or "")))
                except Exception:
                    continue
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 60)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _csv_bytes_from_rows(rows: List[dict]) -> bytes:
    buffer = io.StringIO()
    headers = list(rows[0].keys()) if rows else []
    writer = csv.DictWriter(buffer, fieldnames=headers)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def _manifest_bytes(
    source_filename: str,
    instructions: str,
    label_actions: List[dict],
    action_rows: List[dict],
    counts: Counter,
) -> bytes:
    payload = {
        "generated_at": _now_utc().isoformat(),
        "source_filename": source_filename,
        "instructions": instructions,
        "label_actions": label_actions,
        "counts_by_label": dict(counts),
        "actions": action_rows,
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def _build_outlook_actions_powershell(csv_filename: str) -> str:
    return f"""param(
    [string]$CsvPath = "$(Join-Path $PSScriptRoot '{csv_filename}')"
)

function Resolve-OutlookFolder {{
    param(
        $Namespace,
        [string]$FolderPath
    )

    $clean = ($FolderPath -replace '\\\\', '/').Trim('/')
    if ([string]::IsNullOrWhiteSpace($clean)) {{
        return $null
    }}

    $parts = $clean.Split('/') | Where-Object {{ $_ -and $_.Trim() -ne '' }}
    if ($parts.Count -eq 0) {{
        return $null
    }}

    $root = $Namespace.GetDefaultFolder(6)
    if ($parts[0] -ieq 'Inbox') {{
        if ($parts.Count -eq 1) {{
            return $root
        }}
        $parts = $parts[1..($parts.Count - 1)]
    }}
    $current = $root
    foreach ($part in $parts) {{
        if (-not $part) {{ continue }}
        $next = $null
        foreach ($folder in $current.Folders) {{
            if ($folder.Name -ieq $part) {{
                $next = $folder
                break
            }}
        }}
        if ($null -eq $next) {{
            $next = $current.Folders.Add($part)
        }}
        $current = $next
    }}
    return $current
}}

if (-not (Test-Path $CsvPath)) {{
    throw "CSV de ações não encontrado: $CsvPath"
}}

$outlook = New-Object -ComObject Outlook.Application
$namespace = $outlook.GetNamespace('MAPI')
$rows = Import-Csv -Path $CsvPath

foreach ($row in $rows) {{
    if (-not $row.EntryID) {{ continue }}
    try {{
        $item = $null
        if ($row.StoreID) {{
            $item = $namespace.GetItemFromID($row.EntryID, $row.StoreID)
        }} else {{
            $item = $namespace.GetItemFromID($row.EntryID)
        }}
        if ($null -eq $item) {{ continue }}

        switch (($row.ActionType | ForEach-Object {{ $_.ToLowerInvariant() }})) {{
            'move' {{
                $folder = Resolve-OutlookFolder -Namespace $namespace -FolderPath $row.ActionTarget
                if ($folder) {{
                    $null = $item.Move($folder)
                }}
            }}
            'category' {{
                $item.Categories = $row.ActionTarget
                $item.Save()
            }}
            'flag' {{
                $item.FlagRequest = if ($row.Label) {{ $row.Label }} else {{ 'Follow up' }}
                $interval = '{json.dumps(_FLAG_INTERVAL_BY_TARGET)}' | ConvertFrom-Json
                $target = ($row.ActionTarget | ForEach-Object {{ $_.ToLowerInvariant() }})
                $markValue = if ($interval.PSObject.Properties.Name -contains $target) {{ [int]$interval.$target }} else {{ 0 }}
                $item.MarkAsTask($markValue)
                $item.Save()
            }}
            default {{
                $item.Save()
            }}
        }}
    }} catch {{
        Write-Warning ("Falha a aplicar ação para EntryID {{0}}: {{1}}" -f $row.EntryID, $_.Exception.Message)
    }}
}}
"""


def _build_outlook_draft_powershell(payload: Dict[str, Any], msg_filename: str) -> str:
    payload_b64 = base64.b64encode(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    ).decode("ascii")
    return f"""$draftDir = [string]$env:DBDE_DRAFT_DIR
if ([string]::IsNullOrWhiteSpace($draftDir)) {{
    $draftDir = (Get-Location).Path
}}

$OutputPath = Join-Path $draftDir '{msg_filename}'
$payloadJson = [System.Text.Encoding]::UTF8.GetString([System.Convert]::FromBase64String('{payload_b64}'))
$payload = $payloadJson | ConvertFrom-Json
$outlook = New-Object -ComObject Outlook.Application
$mail = $outlook.CreateItem(0)
$mail.To = [string]$payload.to
$mail.CC = [string]$payload.cc
$mail.BCC = [string]$payload.bcc
$mail.Subject = [string]$payload.subject
if ([string]::IsNullOrWhiteSpace([string]$payload.html_body) -eq $false) {{
    $mail.HTMLBody = [string]$payload.html_body
}} else {{
    $mail.Body = [string]$payload.text_body
}}
foreach ($attachment in $payload.attachments) {{
    if (Test-Path $attachment) {{
        $null = $mail.Attachments.Add($attachment)
    }}
}}
if ([string]::IsNullOrWhiteSpace([string]$OutputPath)) {{
    throw "OutputPath inválido para o draft MSG."
}}
if (([string]$OutputPath).ToLowerInvariant().EndsWith('.msg') -eq $false) {{
    $OutputPath = [string]$OutputPath + '.msg'
}}
if (Test-Path $OutputPath) {{
    Remove-Item -Path $OutputPath -Force
}}
try {{
    $mail.SaveAs($OutputPath, 9)
}} catch {{
    $mail.SaveAs($OutputPath, 3)
}}
$mail.Display()
Write-Host ("Draft MSG criado em: {{0}}" -f $OutputPath)
"""


def _build_outlook_draft_cmd(payload: Dict[str, Any], msg_filename: str) -> str:
    ps_script = _build_outlook_draft_powershell(payload, msg_filename)
    encoded_ps = base64.b64encode(ps_script.encode("utf-16le")).decode("ascii")
    return f"""@echo off
setlocal
set "DBDE_DRAFT_DIR=%~dp0"
powershell.exe -NoLogo -NoProfile -ExecutionPolicy Bypass -EncodedCommand {encoded_ps}
set "EXITCODE=%ERRORLEVEL%"
if not "%EXITCODE%"=="0" (
  echo.
  echo Falha ao criar ou abrir o draft Outlook.
  echo.
  echo Verifica se o Outlook desktop esta instalado e configurado nesta maquina.
  echo Se o problema persistir, abre Windows PowerShell e executa este ficheiro a partir de la para ver o erro.
  pause
)
exit /b %EXITCODE%
"""


async def _store_downloads(files: List[dict], user_sub: str = "") -> List[dict]:
    downloads = []
    for item in files:
        content = item.get("content", b"")
        if not isinstance(content, (bytes, bytearray)) or not content:
            continue
        filename = str(item.get("filename", "") or "download.bin")
        fmt = filename.rsplit(".", 1)[-1].lower() if "." in filename else "bin"
        mime_type = str(item.get("mime_type", "") or "application/octet-stream")
        download_id = await _store_generated_file(
            bytes(content),
            mime_type,
            filename,
            fmt,
            user_sub=str(user_sub or "").strip(),
            scope="prepare_outlook_draft",
        )
        if not download_id:
            continue
        downloads.append(
            {
                "download_id": download_id,
                "endpoint": f"/api/download/{download_id}",
                "filename": filename,
                "format": fmt,
                "mime_type": mime_type,
                "size_bytes": len(content),
                "expires_in_seconds": _GENERATED_FILE_TTL_SECONDS,
                "label": str(item.get("label", "") or "").strip(),
                "description": str(item.get("description", "") or "").strip(),
                "primary": bool(item.get("primary", False)),
            }
        )
    return downloads


async def tool_prepare_outlook_draft(
    subject: str = "",
    body: str = "",
    to: Any = None,
    cc: Any = None,
    bcc: Any = None,
    body_format: str = "html",
    attachments: Optional[List[str]] = None,
    user_sub: str = "",
):
    safe_subject = str(subject or "").strip() or "Draft"
    safe_body = str(body or "").strip()
    to_list = _as_list(to)
    cc_list = _as_list(cc)
    bcc_list = _as_list(bcc)
    attachment_list = _as_list(attachments)

    base_name = _safe_name(safe_subject, "email_draft")
    msg_filename = f"{base_name}.msg"
    payload = {
        "subject": safe_subject,
        "to": "; ".join(to_list),
        "cc": "; ".join(cc_list),
        "bcc": "; ".join(bcc_list),
        "html_body": safe_body if str(body_format or "").lower() == "html" else "",
        "text_body": _html_to_text(safe_body) if str(body_format or "").lower() == "html" else safe_body,
        "attachments": attachment_list,
        "generated_at": _now_utc().isoformat(),
    }
    downloads = await _store_downloads(
        [
            {
                "label": "Gerar .msg e abrir draft no Outlook (.cmd)",
                "description": "Launcher de um clique que cria um ficheiro .msg local e abre o draft no Outlook.",
                "primary": True,
                "filename": f"Open_{base_name}.cmd",
                "mime_type": "text/plain",
                "content": _build_outlook_draft_cmd(payload, msg_filename).encode("utf-8"),
            },
        ],
        user_sub=str(user_sub or "").strip(),
    )
    return {
        "status": "ok",
        "draft_type": "outlook",
        "subject": safe_subject,
        "to": to_list,
        "cc": cc_list,
        "bcc": bcc_list,
        "body_format": "html" if str(body_format or "").lower() == "html" else "text",
        "attachments": attachment_list,
        "summary": (
            "Rascunho preparado para Outlook. "
            "O .cmd gera um ficheiro .msg local e abre o draft no Outlook."
        ),
        "items": [
            {
                "artifact": "draft",
                "subject": safe_subject,
                "to": ", ".join(to_list),
                "body_preview": (_html_to_text(safe_body) or safe_body)[:240],
                "generated_file": msg_filename,
            }
        ],
        "total_count": 1,
        "_auto_file_downloads": downloads,
    }


async def tool_classify_uploaded_emails(
    instructions: str = "",
    conv_id: str = "",
    user_sub: str = "",
    filename: str = "",
    label_actions: Optional[List[dict]] = None,
    batch_size: int = _EMAIL_CLASSIFICATION_DEFAULT_BATCH,
):
    safe_instructions = str(instructions or "").strip()
    if not safe_instructions:
        return {"error": "instructions é obrigatório para classificar emails."}

    try:
        source_filename, source_columns, source_rows = await _load_uploaded_email_table(conv_id, user_sub=user_sub, filename=filename)
    except Exception as e:
        return {"error": str(e)}

    normalized_rows = [
        _canonicalize_email_record(row, str(idx))
        for idx, row in enumerate(source_rows, start=1)
    ]
    label_actions_norm = _normalize_label_actions(label_actions)
    fallback = _fallback_label(label_actions_norm)
    schema = _build_batch_schema(label_actions_norm)

    decisions: Dict[str, dict] = {}
    for chunk in _chunk_list(normalized_rows, int(batch_size or _EMAIL_CLASSIFICATION_DEFAULT_BATCH)):
        prompt = _build_classification_prompt(safe_instructions, label_actions_norm, fallback, chunk)
        parsed = await _classify_email_chunk_with_retry(prompt, schema=schema)
        for entry in parsed.get("decisions", []) if isinstance(parsed, dict) else []:
            row_id = str(entry.get("row_id", "") or "").strip()
            if row_id:
                decisions[row_id] = entry
        for row in chunk:
            decisions.setdefault(
                row["row_id"],
                {
                    "row_id": row["row_id"],
                    "label": fallback,
                    "confidence": 0.0,
                    "reason": "Fallback automático por ausência de decisão estruturada.",
                    "summary": row.get("subject", "")[:140],
                    "requires_manual_review": True,
                },
            )

    output_rows, action_rows, counts = _merge_decisions(source_rows, normalized_rows, decisions, label_actions_norm)
    if not output_rows:
        return {"error": "Não consegui gerar output classificável."}

    timestamp = _now_utc().strftime("%Y%m%d_%H%M%S")
    base_name = _safe_name(source_filename.rsplit(".", 1)[0], "email_actions")
    output_base = f"{base_name}_classified_{timestamp}"
    csv_filename = f"{output_base}.csv"
    downloads = await _store_downloads(
        [
            {
                "label": "Aplicar ações no Outlook (.ps1)",
                "description": "Descarrega e executa localmente para aplicar flags, categorias e movimentos por EntryID.",
                "primary": True,
                "filename": f"Apply_{output_base}.ps1",
                "mime_type": "text/plain",
                "content": _build_outlook_actions_powershell(csv_filename).encode("utf-8"),
            },
            {
                "label": "Excel classificado (.xlsx)",
                "description": "Workbook com Output, Actions e Config.",
                "filename": f"{output_base}.xlsx",
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "content": _workbook_bytes_from_rows(output_rows, action_rows, label_actions_norm, safe_instructions, source_filename),
            },
            {
                "label": "CSV de ações Outlook (.csv)",
                "description": "Ficheiro simples para ingestão local por VBA/PowerShell.",
                "filename": csv_filename,
                "mime_type": "text/csv",
                "content": _csv_bytes_from_rows(action_rows),
            },
            {
                "label": "Manifest JSON (.json)",
                "description": "Critérios, distribuição e ações geradas.",
                "filename": f"{output_base}.json",
                "mime_type": "application/json",
                "content": _manifest_bytes(source_filename, safe_instructions, label_actions_norm, action_rows, counts),
            },
        ]
    )

    preview_items = []
    for row in action_rows[:20]:
        preview_items.append(
            {
                "entry_id": row.get("EntryID", ""),
                "subject": row.get("Subject", ""),
                "from": row.get("From", ""),
                "label": row.get("Label", ""),
                "action_type": row.get("ActionType", ""),
                "action_target": row.get("ActionTarget", ""),
                "confidence": row.get("Confidence", 0),
                "reason": row.get("Reason", ""),
            }
        )

    summary_parts = [f"{label}: {count}" for label, count in counts.most_common()]
    return {
        "status": "ok",
        "source_filename": source_filename,
        "source_columns": source_columns,
        "summary": (
            f"Classifiquei {len(action_rows)} emails de '{source_filename}'. "
            f"Distribuição: {', '.join(summary_parts) if summary_parts else 'sem labels'}."
        ),
        "instructions": safe_instructions,
        "label_actions": label_actions_norm,
        "counts_by_label": dict(counts),
        "items": preview_items,
        "total_count": len(action_rows),
        "_auto_file_downloads": downloads,
    }
